#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
客戶資料 Excel → JSON 轉換腳本 v2
修正：
1. 用欄位標題動態偵測位置，不硬編 index
2. 排序：完工日期 → 分公司（潭→員→海）→ 原客戶編號數字
3. 進件編號：按完工年份分組，跨分公司連續
4. 原客戶編號完整保留（掃描檔對應用）
5. 不截斷任何欄位
"""
import openpyxl
import re
import json
import sys
import os
from datetime import datetime, date

EXCEL_DIR = '/Volumes/UC300-2/客戶資料'
OUTPUT_JSON = '/Users/chifangtang/hswork/database/customer_import_v2.json'
OUTPUT_SQL_CUSTOMERS = '/Users/chifangtang/hswork/database/customer_import_customers_v2.sql'
OUTPUT_SQL_CONTACTS = '/Users/chifangtang/hswork/database/customer_import_contacts_v2.sql'

FILES = [
    {'file': 'ERP最新-客戶資料115年3月26日(潭子).xlsx', 'branch': '潭子', 'order': 1},
    {'file': 'ERP最新-客戶資料115年3月16日(員林).xlsx', 'branch': '員林', 'order': 2},
    {'file': 'ERP最新-客戶資料115年03月11日(海線).xlsx', 'branch': '海線', 'order': 3},
]

# ============================================================
# 欄位標題對照（用標題名稱找位置）
# ============================================================
HEADER_MAP = {
    '客戶編號': 'col_customer_no',
    '客戶名稱': 'col_name',
    '地址': 'col_address',
    '聯絡人/電話': 'col_contact',
    '手機': 'col_mobile1',
    '手機號碼': 'col_mobile2',
    '電話/傳真': 'col_phone_fax',
    '完工日期': 'col_completion',
    '備註(保固期限)': 'col_warranty',
    '付款方式': 'col_payment',
    '業務': 'col_sales',
    '官方LINE': 'col_line',
}


def detect_columns(ws):
    """用標題名稱動態偵測欄位位置"""
    cols = {}
    header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
    for i, cell in enumerate(header_row):
        val = cell.value
        if val and str(val).strip() in HEADER_MAP:
            cols[HEADER_MAP[str(val).strip()]] = i

    # 驗證必要欄位都找到了
    required = ['col_customer_no', 'col_name', 'col_completion', 'col_warranty']
    for r in required:
        if r not in cols:
            raise ValueError(f'找不到必要欄位: {r}')

    return cols


def get_cell_raw(row, col_idx):
    """取得 cell 原始值"""
    if col_idx is None or col_idx >= len(row):
        return None
    return row[col_idx].value


def get_cell_str(row, col_idx):
    """取得 cell 字串值（不截斷）"""
    val = get_cell_raw(row, col_idx)
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val
    s = str(val).strip()
    return s if s else None


def parse_roc_date(val):
    """解析民國年日期或西元年日期，回傳 YYYY-MM-DD 或 None"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')

    s = str(val).strip()
    if not s or s in ('', 'NaN', 'nan', 'None'):
        return None

    # 民國年格式: 114/1/11 or 114/01/11 or 114-1-11
    m = re.match(r'^(\d{2,3})[/\-.](\d{1,2})[/\-.](\d{1,2})$', s)
    if m:
        year = int(m.group(1)) + 1911
        month = int(m.group(2))
        day = int(m.group(3))
        # 驗證日期合法性
        try:
            from calendar import monthrange
            max_day = monthrange(year, month)[1]
            if day > max_day:
                day = max_day  # 修正不合法日期（如 2/29 非閏年）
            return f'{year}-{month:02d}-{day:02d}'
        except (ValueError, OverflowError):
            return None

    # 西元年格式: 2024/1/11
    m2 = re.match(r'^(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})', s)
    if m2:
        year = int(m2.group(1))
        month = int(m2.group(2))
        day = int(m2.group(3))
        try:
            from calendar import monthrange
            max_day = monthrange(year, month)[1]
            if day > max_day:
                day = max_day
            return f'{year}-{month:02d}-{day:02d}'
        except (ValueError, OverflowError):
            return None

    return None


def extract_source_company(name_raw):
    """從客戶名稱提取進件分公司 <XX>"""
    if not name_raw:
        return None, name_raw or ''
    s = str(name_raw).strip()
    m = re.match(r'^<([^>]+)>', s)
    if m:
        source = m.group(1).strip()
        if source in ('禾順 勝鴻', '禾順.勝鴻'):
            source = '禾順'
        clean_name = s[m.end():].strip()
        return source, clean_name
    return None, s


def extract_tax_id(text):
    """從文字中提取統一編號（8位數字）"""
    if not text:
        return None
    matches = re.findall(r'#(\d{8})', str(text))
    return matches[0] if matches else None


def clean_customer_name(name_text):
    """清理客戶名稱：移除統編、發票資訊"""
    if not name_text:
        return ''
    clean = str(name_text)
    # 移除 #XXXXXXXX
    clean = re.sub(r'\s*#\d{7,8}', '', clean)
    # 移除發票相關文字
    clean = re.sub(r'發票\s*[:：].*', '', clean)
    clean = re.sub(r'\(發票.*?\)', '', clean)
    # 只取第一行
    lines = clean.split('\n')
    clean = lines[0].strip()
    clean = clean.rstrip(' \t　')
    return clean


def parse_contacts(col_e_val, col_g_val=None):
    """解析聯絡人欄位"""
    contacts = []
    if not col_e_val:
        return contacts

    text = str(col_e_val).strip()
    lines = [l.strip() for l in text.replace('\r\n', '\n').split('\n') if l.strip()]

    for line in lines:
        contact = {'name': '', 'phone': '', 'role': ''}

        # 提取角色
        role_m = re.search(r'[\(（]([^)）]*窗口[^)）]*)[\)）]', line)
        if role_m:
            contact['role'] = role_m.group(1)
            line = line[:role_m.start()] + line[role_m.end():]

        # 提取手機
        phone_m = re.search(r'(09\d{2}[\-\s]?\d{3}[\-\s]?\d{3})', line)
        if phone_m:
            contact['phone'] = phone_m.group(1).replace(' ', '')
            line = line[:phone_m.start()] + line[phone_m.end():]

        # 提取市話
        if not contact['phone']:
            tel_m = re.search(r'(0\d{1,2}[\-\s]?\d{3,4}[\-\s]?\d{3,4})', line)
            if tel_m:
                contact['phone'] = tel_m.group(1)
                line = line[:tel_m.start()] + line[tel_m.end():]

        # 剩下的是姓名
        name = re.sub(r'[#\d\-\s.、，,]+$', '', line).strip()
        name = re.sub(r'^[.、，,\s]+', '', name)
        if name:
            contact['name'] = name

        if contact['name'] or contact['phone']:
            contacts.append(contact)

    # 補充手機
    if col_g_val and contacts:
        g = str(col_g_val).strip()
        if g and contacts[0]['phone'] == '':
            contacts[0]['phone'] = g

    return contacts


def auto_category(name, address=''):
    """依客戶名稱和地址自動判斷分類"""
    text = (name or '') + ' ' + (address or '')
    if re.search(r'住家|自宅|自住|住宅', name or ''):
        return 'residential'
    if re.search(r'社區|大廈|公寓|管委會|管理委員會', text):
        return 'community'
    if re.search(r'寺|廟|宮|教會|教堂|佛堂|道場|善覺|法師|基督|天主|清真|神壇', text):
        return 'religion'
    if re.search(r'協會|基金會|獅子會|扶輪社|工會|商會|公會|學會|團體|慈濟|紅十字', text):
        return 'association'
    if re.search(r'銀行|郵局|信用合作社|農會|漁會|壽險|保險公司|證券|期貨|保險', text):
        return 'financial'
    if re.search(r'學校|國小|國中|高中|高職|大學|學院|幼兒園|幼稚園|補習班|文教|安親班|托兒所', text):
        return 'school'
    if re.search(r'醫院|診所|醫材|藥局|獸醫|動物醫院|長照|養護|安養|護理|仁愛之家|牙醫|眼科|骨科|中醫|藥房|健康', text):
        return 'hospital'
    if re.search(r'公所|市政|區公所|警察|消防|法院|台電|電力公司|戶政|地政|稅務|鄉公所|鎮公所|監理|國稅|縣政府|市政府|鄉民代表|環境部|環保署|環境保護|水利署|氣象|國防|軍|憲兵|海巡|移民署', text):
        return 'government'
    if re.search(r'台積電|鴻海|台塑|聯電|友達|群創|日月光|中鋼|台化|正新|寶成|統一|味全|台達電|光寶|仁寶|英業達|和碩|緯創|廣達|研華|台灣大哥大|中華電信|遠傳|富邦|國泰|九族|童綜合|京元電子|瑞儀光電|矽品|南亞|台玻|永豐餘|華碩|宏碁|微星|技嘉|聯發科|大立光|可成|鳳凰光學|上銀|捷普|台灣高鐵|中油|中華郵政', text):
        return 'enterprise'
    if re.search(r'建設|建築|營造|開發|建商', text):
        return 'builder'
    if re.search(r'工業|工廠|製造|鑄造|鍛造|鋼鐵|化工|科技股份|精密|機械廠|加工|茶廠|酒廠|食品廠|紡織|塑膠|橡膠|金屬|鋁|銅|鐵工', text):
        return 'industrial'
    if re.search(r'物流|倉儲|運輸|貨運|快遞|宅配|搬家', text):
        return 'logistics'
    if re.search(r'民宿|旅館|旅店|飯店|背包客|旅社|汽車旅館|渡假|度假村|套房|日租|月租|包租|出租|租屋', text):
        return 'hotel'
    if re.search(r'KTV|網咖|電影|遊樂|健身|瑜珈|SPA|按摩|游泳|球場|運動|休閒|娛樂|撞球|保齡球|釣蝦|釣魚|高爾夫|射擊|密室|桌遊|電競|實驗室|玩笑', text):
        return 'leisure'
    if re.search(r'餐廳|小吃|麵店|麵館|麵行|鍋|炸雞|便當|咖啡|早餐|火鍋|燒烤|燒肉|牛排|壽司|拉麵|滷味|雞排|豆漿|飲料|酒|啤酒|肉圓|鹹酥雞|甜不辣|麵包|蛋糕|烘焙|鍋物|食堂|簡餐|豆腐|豆花|冰品|冰店|涮|串燒|居酒屋|披薩|漢堡|排骨|水餃|包子|饅頭|粥|羹|米糕|碗粿|工夫茶|功夫茶|手作茶|手搖|泡沫紅茶|冷飲|茶坊|茶枋|茶棧|茶飲|新井茶|烤肉|羊肉爐|鵝肉|控肉飯|鍋貼|臭豆腐|蚵仔|肉燥|雞肉飯|排骨飯|焢肉|魯肉|牛肉麵|陽春麵|乾麵|炒飯|炒麵|熱炒|快炒|合菜|辦桌|外燴', text):
        return 'food'
    if re.search(r'檳榔|童裝|服飾|美髮|美容|超市|超商|商店|百貨|藥妝|眼鏡|鐘錶|花店|水果|洗衣|修車|輪胎|機車行|汽車|材料行|五金|水電行|印刷|影印|加油站|便利|賣場|寵物|文具|書店|3C|通訊|手機|電信|飾品|珠寶|金飾|銀樓|當鋪|禮品|文創|市集|雜貨|量販|家具|家電|燈飾|窗簾|油漆|裝潢|茶莊|茶葉|茶行|雞肉|豬肉|肉品|肉舖|肉攤|市場|菜市|魚|海鮮|蔬果|米店|雜糧', text):
        return 'shop'
    if re.search(r'股份有限公司|有限公司|企業社|實業|貿易|經銷|批發|保全|企業|公司|對講機|監視|弱電|消防設備|空調|冷氣|電梯', text):
        return 'commercial'
    if re.search(r'商行|行號', text):
        return 'shop'
    return 'residential'


def extract_customer_no_number(original_no):
    """從原客戶編號提取數字部分，用於排序"""
    if not original_no:
        return 999999
    # 找出所有數字
    nums = re.findall(r'\d+', str(original_no))
    if nums:
        # 取最大的數字（通常是主編號）
        return max(int(n) for n in nums)
    return 999999


# ============================================================
# 主程式
# ============================================================
def main():
    all_records = []

    for finfo in FILES:
        filepath = os.path.join(EXCEL_DIR, finfo['file'])
        branch = finfo['branch']
        branch_order = finfo['order']

        print(f'讀取: {branch} ({filepath})')

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb['施工客戶']

        # 動態偵測欄位位置
        cols = detect_columns(ws)
        print(f'  欄位偵測結果:')
        for k, v in sorted(cols.items(), key=lambda x: x[1]):
            print(f'    [{v}] {k}')

        row_count = 0
        skip_count = 0

        for row in ws.iter_rows(min_row=2, values_only=False):
            col_a = get_cell_str(row, cols.get('col_customer_no'))
            col_b = get_cell_str(row, cols.get('col_name'))

            # 跳過空行
            if not col_b and not col_a:
                skip_count += 1
                continue

            row_count += 1

            # 地址
            col_d = get_cell_str(row, cols.get('col_address'))
            # 聯絡人/電話
            col_e = get_cell_str(row, cols.get('col_contact'))
            # 手機
            col_f = get_cell_str(row, cols.get('col_mobile1'))
            # 手機號碼
            col_g = get_cell_str(row, cols.get('col_mobile2'))
            # 電話/傳真
            col_h = get_cell_str(row, cols.get('col_phone_fax'))
            # 完工日期（取原始值，不轉字串）
            col_completion_raw = get_cell_raw(row, cols.get('col_completion'))
            # 保固
            col_warranty = get_cell_str(row, cols.get('col_warranty'))
            # 付款
            col_payment = get_cell_str(row, cols.get('col_payment'))
            # 業務
            col_sales = get_cell_str(row, cols.get('col_sales'))
            # LINE
            col_line = get_cell_str(row, cols.get('col_line'))

            # 解析名稱
            name_raw = str(col_b) if col_b else ''
            source_company, name_after = extract_source_company(name_raw)

            # 統編
            tax_id = extract_tax_id(name_after)
            if not tax_id and col_d:
                tax_id = extract_tax_id(str(col_d))

            # 清理客戶名稱
            customer_name = clean_customer_name(name_after)

            # 完工日期
            completion_date = parse_roc_date(col_completion_raw)

            # 保固
            warranty_date = None
            warranty_note = None
            if col_warranty:
                wd = parse_roc_date(col_warranty)
                if wd:
                    warranty_date = wd
                else:
                    warranty_note = str(col_warranty).strip()

            # 電話/傳真
            phone = ''
            fax = ''
            if col_h:
                h_str = str(col_h)
                if h_str.upper().startswith('F:') or h_str.upper().startswith('FAX:'):
                    fax = re.sub(r'^(F|FAX)\s*:\s*', '', h_str, flags=re.IGNORECASE).strip()
                elif 'FAX' in h_str.upper() or 'F:' in h_str.upper() or 'F：' in h_str:
                    fax = h_str
                else:
                    phone = h_str

            # 聯絡人
            contacts = parse_contacts(col_e, col_g)
            primary_contact = contacts[0]['name'] if contacts else ''
            primary_phone = contacts[0]['phone'] if contacts else ''

            # 手機
            mobile = str(col_g).strip() if col_g else ''

            # 最終電話：如果電話/傳真欄是空的，用聯絡人裡提取的
            if not phone:
                phone = primary_phone

            # 分類
            category = auto_category(customer_name, str(col_d) if col_d else '')

            # 原客戶編號
            original_no = str(col_a).strip() if col_a else ''

            record = {
                'original_customer_no': original_no,
                'source_company': source_company,
                'customer_name': customer_name,
                'category': category,
                'tax_id': tax_id,
                'site_address': str(col_d) if col_d else None,
                'contact_person': primary_contact,
                'phone': phone,
                'mobile': mobile,
                'fax': fax,
                'contacts': contacts,
                'completion_date': completion_date,
                'warranty_date': warranty_date,
                'warranty_note': warranty_note,
                'payment_info': str(col_payment) if col_payment else None,
                'salesperson_name': str(col_sales) if col_sales else None,
                'line_official': str(col_line) if col_line else None,
                'source_branch': branch,
                'branch_order': branch_order,
                '_sort_no': extract_customer_no_number(original_no),
            }
            all_records.append(record)

        print(f'  讀取: {row_count} 筆, 跳過空行: {skip_count}')
        wb.close()

    print(f'\n總計: {len(all_records)} 筆')

    # ============================================================
    # 排序：完工日期 → 分公司順序 → 原客戶編號數字
    # ============================================================
    print('\n排序中...')

    def sort_key(r):
        d = r.get('completion_date') or '9999-12-31'
        branch_ord = r.get('branch_order', 9)
        no = r.get('_sort_no', 999999)
        return (d, branch_ord, no)

    all_records.sort(key=sort_key)

    # ============================================================
    # 編號
    # ============================================================
    print('產生編號...')

    # 進件編號：按完工年份分組，跨分公司連續
    year_counters = {}
    for r in all_records:
        d = r.get('completion_date')
        year = d[:4] if d else '2026'
        if year not in year_counters:
            year_counters[year] = 0
        year_counters[year] += 1
        r['case_number'] = f'{year}-{year_counters[year]:04d}'

    # 客戶編號：A-六位流水號
    for idx, r in enumerate(all_records):
        r['customer_no'] = f'A-{idx + 1:06d}'

    # 移除排序用的暫存欄位
    for r in all_records:
        r.pop('branch_order', None)
        r.pop('_sort_no', None)

    # ============================================================
    # 關聯群組
    # ============================================================
    print('建立關聯群組...')
    group_id = 0
    tax_groups = {}
    name_count = {}

    # 第一輪：統編
    for r in all_records:
        if r.get('tax_id'):
            tid = r['tax_id']
            if tid not in tax_groups:
                group_id += 1
                tax_groups[tid] = group_id
            r['related_group_id'] = tax_groups[tid]

    # 第二輪：同名
    for r in all_records:
        if r.get('related_group_id'):
            continue
        cname = (r.get('customer_name') or '').strip()
        if not cname:
            continue
        if cname not in name_count:
            name_count[cname] = []
        name_count[cname].append(r)

    for cname, recs in name_count.items():
        if len(recs) > 1:
            group_id += 1
            for r in recs:
                r['related_group_id'] = group_id

    grouped = sum(1 for r in all_records if r.get('related_group_id'))
    print(f'  關聯群組: {group_id} 組, 涉及 {grouped} 筆')

    # ============================================================
    # 統計
    # ============================================================
    with_date = sum(1 for r in all_records if r.get('completion_date'))
    with_tax = sum(1 for r in all_records if r.get('tax_id'))
    with_contacts = sum(1 for r in all_records if r.get('contacts'))

    print(f'\n=== 統計 ===')
    print(f'  總筆數: {len(all_records)}')
    print(f'  有完工日期: {with_date}')
    print(f'  有統編: {with_tax}')
    print(f'  有聯絡人: {with_contacts}')

    # 各分公司
    br = {}
    for r in all_records:
        b = r.get('source_branch', '?')
        br[b] = br.get(b, 0) + 1
    print(f'\n來源分公司:')
    for k, v in br.items():
        print(f'  {k}: {v}')

    # 年度分布
    print(f'\n年度分布:')
    for year in sorted(year_counters.keys()):
        print(f'  {year}: {year_counters[year]}')

    # 前5筆
    print(f'\n=== 前5筆 ===')
    for r in all_records[:5]:
        print(f'  {r["customer_no"]} | {r["case_number"]} | {r["original_customer_no"]} | <{r.get("source_company","")}>  {r["customer_name"]} | {r["completion_date"]} | {r["source_branch"]}')

    # 各分公司邊界
    print(f'\n=== 分公司邊界 ===')
    prev_branch = None
    for i, r in enumerate(all_records):
        if r['source_branch'] != prev_branch:
            print(f'  [{i}] {r["customer_no"]} | {r["source_branch"]} 開始 | {r["completion_date"]} | {r["original_customer_no"]}')
            prev_branch = r['source_branch']

    # ============================================================
    # 輸出 JSON
    # ============================================================
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(all_records, f, ensure_ascii=False, indent=None)
    print(f'\nJSON: {OUTPUT_JSON}')
    print(f'  大小: {os.path.getsize(OUTPUT_JSON):,} bytes')

    # ============================================================
    # 輸出 SQL — customers
    # ============================================================
    print(f'\n產生 SQL...')

    def sql_escape(val):
        if val is None:
            return 'NULL'
        s = str(val)
        s = s.replace('\\', '\\\\').replace("'", "\\'")
        return f"'{s}'"

    with open(OUTPUT_SQL_CUSTOMERS, 'w', encoding='utf-8') as f:
        f.write('SET NAMES utf8mb4;\n')
        f.write('SET FOREIGN_KEY_CHECKS=0;\n\n')
        f.write('TRUNCATE TABLE customer_contacts;\n')
        f.write('DELETE FROM customer_groups;\n')
        f.write('ALTER TABLE customer_groups AUTO_INCREMENT = 1;\n')
        f.write('UPDATE cases SET customer_id = NULL, customer_no = NULL;\n')
        f.write('DELETE FROM customers;\n')
        f.write('ALTER TABLE customers AUTO_INCREMENT = 1;\n\n')

        f.write('INSERT INTO customers (customer_no,case_number,name,category,source_company,original_customer_no,related_group_id,contact_person,phone,mobile,tax_id,site_address,completion_date,warranty_date,warranty_note,payment_info,payment_terms,salesperson_name,sales_id,line_official,source_branch,import_source,is_active,created_at) VALUES\n')

        lines = []
        for r in all_records:
            vals = [
                sql_escape(r['customer_no']),
                sql_escape(r['case_number']),
                sql_escape(r['customer_name']),
                sql_escape(r['category']),
                sql_escape(r.get('source_company')),
                sql_escape(r['original_customer_no']),
                str(r.get('related_group_id', 'NULL')) if r.get('related_group_id') else 'NULL',
                sql_escape(r.get('contact_person')),
                sql_escape(r.get('phone')),
                sql_escape(r.get('mobile')),
                sql_escape(r.get('tax_id')),
                sql_escape(r.get('site_address')),
                sql_escape(r.get('completion_date')),
                sql_escape(r.get('warranty_date')),
                sql_escape(r.get('warranty_note')),
                sql_escape(r.get('payment_info')),
                'NULL',  # payment_terms
                sql_escape(r.get('salesperson_name')),
                'NULL',  # sales_id
                sql_escape(r.get('line_official')),
                sql_escape(r.get('source_branch')),
                "'excel_import'",
                '1',
                'NOW()',
            ]
            lines.append('(' + ','.join(vals) + ')')

        f.write(',\n'.join(lines))
        f.write(';\n\nSET FOREIGN_KEY_CHECKS=1;\n')

    print(f'SQL (customers): {OUTPUT_SQL_CUSTOMERS}')

    # ============================================================
    # 輸出 SQL — contacts
    # ============================================================
    with open(OUTPUT_SQL_CONTACTS, 'w', encoding='utf-8') as f:
        f.write('SET NAMES utf8mb4;\n\n')
        f.write('TRUNCATE TABLE customer_contacts;\n\n')
        f.write('INSERT INTO customer_contacts (customer_id,contact_name,phone,role) VALUES\n')

        lines = []
        for idx, r in enumerate(all_records):
            cid = idx + 1  # customer ID = 順序 +1 (因為 AUTO_INCREMENT 從 1 開始)
            if r.get('contacts'):
                for c in r['contacts']:
                    if c.get('name') or c.get('phone'):
                        lines.append(f"({cid},{sql_escape(c.get('name',''))},{sql_escape(c.get('phone',''))},{sql_escape(c.get('role',''))})")

        f.write(',\n'.join(lines))
        f.write(';\n')

    print(f'SQL (contacts): {OUTPUT_SQL_CONTACTS}')
    print(f'\n完成！')


if __name__ == '__main__':
    main()
