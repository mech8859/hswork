#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 原始檔 vs DB 源頭比對
直接讀 Excel 欄位，和 DB 的 customers 表逐筆比較
確認匯入過程（Excel→Python→JSON→SQL→DB）有沒有弄錯
"""
import openpyxl
import re
import json
import sys
import os
from datetime import datetime, date

EXCEL_DIR = '/Volumes/UC300-2/客戶資料'
FILES = [
    {'file': 'ERP最新-客戶資料115年3月26日(潭子).xlsx', 'branch': '潭子'},
    {'file': 'ERP最新-客戶資料115年3月16日(員林).xlsx', 'branch': '員林'},
    {'file': 'ERP最新-客戶資料115年03月11日(海線).xlsx', 'branch': '海線'},
]

# 讀取 JSON（已匯入 DB 的中間檔），用 customer_no 索引
JSON_FILE = '/Users/chifangtang/hswork/database/customer_import.json'

def get_cell(row, col_idx):
    if col_idx < len(row):
        v = row[col_idx].value
        if v is not None:
            return str(v).strip() if not isinstance(v, (datetime, date)) else v
    return None

def parse_roc_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s or s in ('', 'NaN', 'nan'):
        return None
    m = re.match(r'^(\d{2,3})[/\-.](\d{1,2})(?:[/\-.](\d{1,2}))?$', s)
    if m:
        year = int(m.group(1)) + 1911
        month = int(m.group(2))
        day = int(m.group(3)) if m.group(3) else 1
        return f'{year}-{month:02d}-{day:02d}'
    m2 = re.match(r'^(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})', s)
    if m2:
        return f'{m2.group(1)}-{int(m2.group(2)):02d}-{int(m2.group(3)):02d}'
    return None

def extract_source_company(name_raw):
    if not name_raw:
        return None, name_raw
    m = re.match(r'^<([^>]+)>', name_raw)
    if m:
        source = m.group(1).strip()
        if source in ('禾順 勝鴻', '禾順.勝鴻'):
            source = '禾順'
        clean_name = name_raw[m.end():].strip()
        return source, clean_name
    return None, name_raw

def extract_tax_id(text):
    if not text:
        return None
    matches = re.findall(r'#(\d{7,8})', text)
    for m in matches:
        if len(m) == 8:
            return m
    return None

def extract_invoice_info(name_text):
    if not name_text:
        return None, name_text
    tax_id = extract_tax_id(name_text)
    clean = name_text
    clean = re.sub(r'#\d{7,8}', '', clean)
    clean = re.sub(r'發票\s*[:：].*', '', clean)
    clean = re.sub(r'\(發票.*?\)', '', clean)
    lines = clean.split('\n')
    clean = lines[0].strip()
    clean = clean.rstrip(' \t　')
    return tax_id, clean

def main():
    # 載入 JSON（中間檔），用來找 customer_no 對應
    print("載入 JSON 中間檔...")
    with open(JSON_FILE, 'r', encoding='utf-8') as f:
        json_records = json.load(f)

    # 建立索引：用 (branch, original_customer_no, customer_name) 去對應 customer_no
    # 但更可靠的是用順序：JSON 是按完工日期排序再編號的，Excel 是原始順序
    # 所以我們用 customer_no 從 JSON 找回對應的原始 Excel 資料

    # 先建立 JSON 索引 by customer_no
    json_by_no = {}
    for r in json_records:
        json_by_no[r['customer_no']] = r

    print(f"JSON 共 {len(json_records)} 筆")

    # 讀取所有 Excel 原始行
    print("\n讀取 Excel 原始檔...")
    excel_rows = []  # [{branch, row_num, raw_fields...}]

    for finfo in FILES:
        filepath = os.path.join(EXCEL_DIR, finfo['file'])
        branch = finfo['branch']
        print(f"  {branch}: {filepath}")

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb['施工客戶']

        # 偵測欄位
        headers = [c.value for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
        completion_idx = None
        for i, h in enumerate(headers):
            if h and '完工' in str(h):
                completion_idx = i
                break
        if completion_idx is None:
            completion_idx = 9
        warranty_idx = completion_idx + 1
        payment_idx = warranty_idx + 1
        sales_idx = payment_idx + 1
        line_idx = sales_idx + 1
        phone_fax_idx = completion_idx - 1

        print(f"    欄位: 完工={completion_idx}, 保固={warranty_idx}, 電話傳真={phone_fax_idx}")

        row_num = 1
        for row in ws.iter_rows(min_row=2, max_col=max(line_idx+1, 14), values_only=False):
            row_num += 1
            col_a = get_cell(row, 0)  # 原客戶編號
            col_b = get_cell(row, 1)  # 客戶名稱（含<XX>前綴）
            col_d = get_cell(row, 3)  # 地址
            col_e = get_cell(row, 4)  # 聯絡人/電話
            col_g = get_cell(row, 6)  # 手機號碼
            col_h = get_cell(row, phone_fax_idx)  # 電話/傳真
            col_j_raw = row[completion_idx].value if completion_idx < len(row) else None
            col_k = get_cell(row, warranty_idx)
            col_l = get_cell(row, payment_idx)
            col_m = get_cell(row, sales_idx)
            col_n = get_cell(row, line_idx)

            if not col_b and not col_a:
                continue

            # 解析名稱
            source_company, name_after = extract_source_company(col_b or '')
            tax_id, customer_name = extract_invoice_info(name_after)
            if not tax_id and col_d:
                tax_id = extract_tax_id(str(col_d))

            # 解析日期
            completion_date = parse_roc_date(col_j_raw)
            warranty_date = None
            warranty_note = None
            if col_k:
                wd = parse_roc_date(col_k)
                if wd:
                    warranty_date = wd
                else:
                    warranty_note = str(col_k).strip()

            # 電話
            phone = ''
            fax = ''
            if col_h:
                h_str = str(col_h)
                if 'F:' in h_str or 'f:' in h_str:
                    fax = h_str.replace('F:', '').replace('f:', '').strip()
                else:
                    phone = h_str

            excel_rows.append({
                'branch': branch,
                'row_num': row_num,
                'original_customer_no': str(col_a).strip() if col_a else '',
                'raw_name': col_b,
                'customer_name': customer_name,
                'source_company': source_company,
                'tax_id': tax_id,
                'site_address': col_d,
                'contact_person_raw': col_e,  # 原始聯絡人欄
                'mobile_raw': col_g,  # 原始手機欄
                'phone_fax_raw': col_h,  # 原始電話/傳真欄
                'phone': phone,
                'mobile': str(col_g).strip() if col_g else '',
                'completion_date': completion_date,
                'completion_date_raw': str(col_j_raw) if col_j_raw else '',
                'warranty_date': warranty_date,
                'warranty_date_raw': str(col_k) if col_k else '',
                'warranty_note': warranty_note,
                'payment_info': col_l,
                'salesperson_name': col_m,
                'line_official': col_n,
            })

        print(f"    讀取 {row_num - 1} 行")
        wb.close()

    print(f"\nExcel 總計: {len(excel_rows)} 筆")

    # 現在要建立 Excel 行 → customer_no 的對應
    # JSON 是把 Excel 資料按 completion_date 排序後，從 A-000001 開始編號
    # 所以我們用同樣的排序方式來對應

    def sort_key(r):
        d = r.get('completion_date')
        if d:
            return d
        return '9999-12-31'

    excel_rows.sort(key=sort_key)

    # 配對 customer_no
    if len(excel_rows) != len(json_records):
        print(f"\n警告: Excel {len(excel_rows)} 筆 vs JSON {len(json_records)} 筆，數量不符！")

    for idx, er in enumerate(excel_rows):
        er['customer_no'] = f'A-{idx + 1:06d}'

    # 驗證配對正確性：用 original_customer_no + customer_name 確認
    mismatch_count = 0
    for er in excel_rows[:20]:
        cno = er['customer_no']
        jr = json_by_no.get(cno)
        if jr:
            if er['customer_name'] != jr['customer_name']:
                mismatch_count += 1

    if mismatch_count > 5:
        print(f"\n嚴重警告: 配對驗證失敗！前20筆有 {mismatch_count} 筆客戶名不一致")
        print("可能是排序邏輯不同，需要換配對方式")

    # ====================================
    # 開始比對 Excel vs JSON（找出 Python 轉換錯誤）
    # ====================================
    print("\n" + "=" * 60)
    print("Phase 1: Excel vs JSON（Python 轉換是否正確）")
    print("=" * 60)

    fields_to_compare = [
        ('customer_name', '客戶名稱'),
        ('source_company', '進件公司'),
        ('original_customer_no', '原客戶編號'),
        ('tax_id', '統編'),
        ('site_address', '地址'),
        ('phone', '電話'),
        ('mobile', '手機'),
        ('completion_date', '完工日期'),
        ('warranty_date', '保固日期'),
        ('warranty_note', '保固備註'),
        ('payment_info', '付款方式'),
        ('salesperson_name', '業務'),
        ('line_official', '官方LINE'),
    ]

    field_diffs = {f[0]: [] for f in fields_to_compare}

    for er in excel_rows:
        cno = er['customer_no']
        jr = json_by_no.get(cno)
        if not jr:
            continue

        for field, label in fields_to_compare:
            ev = er.get(field)
            jv = jr.get(field)

            # 正規化
            ev_str = str(ev).strip() if ev else ''
            jv_str = str(jv).strip() if jv else ''

            if ev_str == 'None':
                ev_str = ''
            if jv_str == 'None':
                jv_str = ''

            if ev_str != jv_str:
                if len(field_diffs[field]) < 10:
                    field_diffs[field].append({
                        'customer_no': cno,
                        'branch': er['branch'],
                        'row': er['row_num'],
                        'excel': ev_str,
                        'json': jv_str,
                        'raw': er.get(field + '_raw', ''),
                    })

    for field, label in fields_to_compare:
        diffs = field_diffs[field]
        total = len([er for er in excel_rows if er['customer_no'] in json_by_no])
        diff_count = sum(1 for er in excel_rows
                        if er['customer_no'] in json_by_no
                        and str(er.get(field) or '').strip().replace('None','') != str(json_by_no[er['customer_no']].get(field) or '').strip().replace('None',''))

        status = "OK" if diff_count == 0 else f"差異 {diff_count} 筆"
        print(f"\n  {label} ({field}): {status}")

        if diffs:
            for d in diffs[:5]:
                print(f"    {d['customer_no']} [{d['branch']} R{d['row']}]")
                print(f"      Excel: {d['excel'][:80]}")
                print(f"      JSON:  {d['json'][:80]}")

    # ====================================
    # Phase 2: 列出 Excel 原始值 vs DB 應該有的值
    # 直接輸出報告供檢視
    # ====================================
    print("\n" + "=" * 60)
    print("Phase 2: Excel 原始欄位抽樣（前10、中10、後10）")
    print("=" * 60)

    samples = excel_rows[:10] + excel_rows[len(excel_rows)//2:len(excel_rows)//2+10] + excel_rows[-10:]

    for er in samples:
        cno = er['customer_no']
        jr = json_by_no.get(cno, {})
        print(f"\n--- {cno} [{er['branch']} R{er['row_num']}] ---")
        print(f"  原始名稱: {er['raw_name']}")
        print(f"  解析名稱: {er['customer_name']}")
        print(f"  JSON名稱: {jr.get('customer_name','')}")
        print(f"  地址: {(er['site_address'] or '')[:60]}")
        print(f"  完工(原始): {er['completion_date_raw']} → 解析: {er['completion_date']}")
        print(f"  保固(原始): {er['warranty_date_raw']} → 解析: {er['warranty_date']} / 備註: {er['warranty_note']}")
        print(f"  電話/傳真(原始): {er['phone_fax_raw']}")
        print(f"  聯絡人(原始): {(er['contact_person_raw'] or '')[:60]}")
        print(f"  手機(原始): {er['mobile_raw']}")
        print(f"  業務: {er['salesperson_name']}")

    print("\n\n完成！")

if __name__ == '__main__':
    main()
