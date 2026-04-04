#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
客戶資料 Excel → JSON 轉換腳本
讀取三個分公司 Excel，解析欄位，產生編號，建立關聯，輸出 JSON
"""
import openpyxl
import re
import json
import sys
from datetime import datetime, date

EXCEL_DIR = '/Volumes/UC300-2/客戶資料'
OUTPUT_FILE = '/Users/chifangtang/hswork/database/customer_import.json'

FILES = [
    {'file': 'ERP最新-客戶資料115年3月26日(潭子).xlsx', 'branch': '潭子'},
    {'file': 'ERP最新-客戶資料115年3月16日(員林).xlsx', 'branch': '員林'},
    {'file': 'ERP最新-客戶資料115年03月11日(海線).xlsx', 'branch': '海線'},
]

def parse_roc_date(val):
    """解析民國年日期或西元年日期"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s or s in ('', 'NaN', 'nan'):
        return None
    # 民國年格式: 114/1/11 or 114/01/11
    m = re.match(r'^(\d{2,3})[/\-.](\d{1,2})(?:[/\-.](\d{1,2}))?$', s)
    if m:
        year = int(m.group(1)) + 1911
        month = int(m.group(2))
        day = int(m.group(3)) if m.group(3) else 1
        try:
            return f'{year}-{month:02d}-{day:02d}'
        except:
            return None
    # 西元年格式
    m2 = re.match(r'^(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})', s)
    if m2:
        return f'{m2.group(1)}-{int(m2.group(2)):02d}-{int(m2.group(3)):02d}'
    return None

def extract_source_company(name_raw):
    """從客戶名稱提取進件分公司 <XX>"""
    if not name_raw:
        return None, name_raw
    m = re.match(r'^<([^>]+)>', name_raw)
    if m:
        source = m.group(1).strip()
        # 統一名稱
        if source in ('禾順 勝鴻', '禾順.勝鴻'):
            source = '禾順'
        clean_name = name_raw[m.end():].strip()
        return source, clean_name
    return None, name_raw

def extract_tax_id(text):
    """從文字中提取統一編號（8位數字）"""
    if not text:
        return None
    matches = re.findall(r'#(\d{7,8})', text)
    for m in matches:
        if len(m) == 8:
            return m
    return None

def extract_invoice_info(name_text):
    """從客戶名稱欄提取發票資訊"""
    if not name_text:
        return None, None, name_text
    tax_id = extract_tax_id(name_text)
    invoice_title = None
    # 提取發票抬頭（通常在統編前的公司名就是）
    # 移除統編和發票地址
    clean = name_text
    # 移除 #XXXXXXXX
    clean = re.sub(r'#\d{7,8}', '', clean)
    # 移除發票相關文字
    clean = re.sub(r'發票\s*[:：].*', '', clean)
    clean = re.sub(r'\(發票.*?\)', '', clean)
    # 移除換行後的內容（通常是發票資訊）
    lines = clean.split('\n')
    clean = lines[0].strip()
    # 移除尾部空白和特殊字元
    clean = clean.rstrip(' \t　')
    return tax_id, invoice_title, clean

def parse_contacts(col_e, col_f=None, col_g=None):
    """解析聯絡人欄位，可能有多位"""
    contacts = []
    if not col_e:
        return contacts

    text = str(col_e).strip()
    # 以換行分隔多位聯絡人
    lines = [l.strip() for l in text.replace('\r\n', '\n').split('\n') if l.strip()]

    for line in lines:
        contact = {'name': '', 'phone': '', 'role': ''}
        # 嘗試提取姓名和電話
        # 常見格式: "林先生 0910-234-352", "採購#1312彭碧瑩", "(報價窗口)"

        # 提取角色 (XXX窗口)
        role_m = re.search(r'[\(（]([^)）]+窗口[^)）]*)[\)）]', line)
        if role_m:
            contact['role'] = role_m.group(1)
            line = line[:role_m.start()] + line[role_m.end():]

        # 提取手機號碼
        phone_m = re.search(r'(09\d{2}[\-\s]?\d{3}[\-\s]?\d{3})', line)
        if phone_m:
            contact['phone'] = phone_m.group(1).replace(' ', '')
            line = line[:phone_m.start()] + line[phone_m.end():]

        # 提取市話
        tel_m = re.search(r'(0\d{1,2}[\-\s]?\d{3,4}[\-\s]?\d{3,4})', line)
        if tel_m and not contact['phone']:
            contact['phone'] = tel_m.group(1)
            line = line[:tel_m.start()] + line[tel_m.end():]

        # 剩下的就是姓名
        name = re.sub(r'[#\d\-\s.、，,]+$', '', line).strip()
        name = re.sub(r'^[.、，,\s]+', '', name)
        if name:
            contact['name'] = name

        if contact['name'] or contact['phone']:
            contacts.append(contact)

    # 補充第一位聯絡人的手機（從 col_g）
    if col_g and contacts:
        g = str(col_g).strip()
        if g and contacts[0]['phone'] == '':
            contacts[0]['phone'] = g

    return contacts

def auto_category(name, address=''):
    """依客戶名稱和地址自動判斷分類"""
    text = (name or '') + ' ' + (address or '')
    # 有「住家」明確標示的優先歸個人
    if re.search(r'住家|自宅|自住|住宅', name or ''):
        return 'residential'
    # 社區／管委會
    if re.search(r'社區|大廈|公寓|管委會|管理委員會', text):
        return 'community'
    # 宗教
    if re.search(r'寺|廟|宮|教會|教堂|佛堂|道場|善覺|法師|基督|天主|清真|神壇', text):
        return 'religion'
    # 協會／團體
    if re.search(r'協會|基金會|獅子會|扶輪社|工會|商會|公會|學會|團體|慈濟|紅十字', text):
        return 'association'
    # 金融／保險
    if re.search(r'銀行|郵局|信用合作社|農會|漁會|壽險|保險公司|證券|期貨|保險', text):
        return 'financial'
    # 教育
    if re.search(r'學校|國小|國中|高中|高職|大學|學院|幼兒園|幼稚園|補習班|文教|安親班|托兒所', text):
        return 'school'
    # 醫療／健康照護
    if re.search(r'醫院|診所|醫材|藥局|獸醫|動物醫院|長照|養護|安養|護理|仁愛之家|牙醫|眼科|骨科|中醫|藥房|健康', text):
        return 'hospital'
    # 機關／政府
    if re.search(r'公所|市政|區公所|警察|消防|法院|台電|電力公司|戶政|地政|稅務|鄉公所|鎮公所|監理|國稅|縣政府|市政府|鄉民代表|環境部|環保署|環境保護|水利署|氣象|國防|軍|憲兵|海巡|移民署', text):
        return 'government'
    # 上市櫃企業
    if re.search(r'台積電|鴻海|台塑|聯電|友達|群創|日月光|中鋼|台化|正新|寶成|統一|味全|台達電|光寶|仁寶|英業達|和碩|緯創|廣達|研華|台灣大哥大|中華電信|遠傳|富邦|國泰|九族|童綜合|京元電子|瑞儀光電|矽品|南亞|台玻|永豐餘|華碩|宏碁|微星|技嘉|聯發科|大立光|可成|鳳凰光學|上銀|捷普|台灣高鐵|中油|中華郵政', text):
        return 'enterprise'
    # 建設／營造
    if re.search(r'建設|建築|營造|開發|建商', text):
        return 'builder'
    # 製造／工廠
    if re.search(r'工業|工廠|製造|鑄造|鍛造|鋼鐵|化工|科技股份|精密|機械廠|加工|茶廠|酒廠|食品廠|紡織|塑膠|橡膠|金屬|鋁|銅|鐵工', text):
        return 'industrial'
    # 物流／倉儲
    if re.search(r'物流|倉儲|運輸|貨運|快遞|宅配|搬家', text):
        return 'logistics'
    # 旅宿業
    if re.search(r'民宿|旅館|旅店|飯店|背包客|旅社|汽車旅館|渡假|度假村|套房|日租|月租|包租|出租|租屋', text):
        return 'hotel'
    # 休閒娛樂
    if re.search(r'KTV|網咖|電影|遊樂|健身|瑜珈|SPA|按摩|游泳|球場|運動|休閒|娛樂|撞球|保齡球|釣蝦|釣魚|高爾夫|射擊|密室|桌遊|電競|實驗室|玩笑', text):
        return 'leisure'
    # 餐飲業
    if re.search(r'餐廳|小吃|麵店|麵館|麵行|鍋|炸雞|便當|咖啡|早餐|火鍋|燒烤|燒肉|牛排|壽司|拉麵|滷味|雞排|豆漿|飲料|酒|啤酒|肉圓|鹹酥雞|甜不辣|麵包|蛋糕|烘焙|鍋物|食堂|簡餐|豆腐|豆花|冰品|冰店|涮|串燒|居酒屋|披薩|漢堡|排骨|水餃|包子|饅頭|粥|羹|米糕|碗粿|工夫茶|功夫茶|手作茶|手搖|泡沫紅茶|冷飲|茶坊|茶枋|茶棧|茶飲|新井茶|烤肉|羊肉爐|鵝肉|控肉飯|鍋貼|臭豆腐|蚵仔|肉燥|雞肉飯|排骨飯|焢肉|魯肉|牛肉麵|陽春麵|乾麵|炒飯|炒麵|熱炒|快炒|合菜|辦桌|外燴', text):
        return 'food'
    # 零售／店面
    if re.search(r'檳榔|童裝|服飾|美髮|美容|超市|超商|商店|百貨|藥妝|眼鏡|鐘錶|花店|水果|洗衣|修車|輪胎|機車行|汽車|材料行|五金|水電行|印刷|影印|加油站|便利|賣場|寵物|文具|書店|3C|通訊|手機|電信|飾品|珠寶|金飾|銀樓|當鋪|禮品|文創|市集|雜貨|量販|家具|家電|燈飾|窗簾|油漆|裝潢|茶莊|茶葉|茶行|雞肉|豬肉|肉品|肉舖|肉攤|市場|菜市|魚|海鮮|蔬果|米店|雜糧', text):
        return 'shop'
    # 一般公司／企業（有公司字樣但無法判斷特定行業的）
    if re.search(r'股份有限公司|有限公司|企業社|實業|貿易|經銷|批發|保全|企業|公司|對講機|監視|弱電|消防設備|空調|冷氣|電梯', text):
        return 'commercial'
    if re.search(r'商行|行號', text):
        return 'shop'
    # 預設個人／住戶
    return 'residential'

CATEGORY_LABELS = {
    'residential': '個人／住戶',
    'food': '餐飲業',
    'shop': '零售／店面',
    'hospital': '醫療／健康照護',
    'school': '教育',
    'religion': '宗教',
    'leisure': '休閒娛樂',
    'hotel': '旅宿業',
    'financial': '金融／保險',
    'industrial': '製造／工廠',
    'builder': '建設／營造',
    'logistics': '物流／倉儲',
    'community': '社區／管委會',
    'government': '機關／政府',
    'commercial': '一般公司／企業',
    'enterprise': '上市櫃企業',
    'association': '協會／團體',
}

def parse_customer_no(raw):
    """解析原客戶編號"""
    if not raw:
        return ''
    return str(raw).strip()

def get_cell(row, col_idx):
    """安全取得 cell 值"""
    if col_idx < len(row):
        v = row[col_idx].value
        if v is not None:
            return str(v).strip() if not isinstance(v, (datetime, date)) else v
    return None

def process_excel(filepath, branch):
    """處理單一 Excel 檔案"""
    print(f'Processing: {filepath} ({branch})')
    wb = openpyxl.load_workbook(filepath, data_only=True)

    records = []

    # 主要處理「施工客戶」sheet
    if '施工客戶' not in wb.sheetnames:
        print(f'  Warning: 施工客戶 sheet not found!')
        return records

    ws = wb['施工客戶']
    total = ws.max_row - 1  # 扣掉標題
    print(f'  Total rows: {total}')

    # 偵測欄位位置：潭子電話/傳真在[8]完工在[9]，員林/海線電話/傳真在[7]完工在[8]
    headers = [c.value for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
    completion_idx = None
    for i, h in enumerate(headers):
        if h and '完工' in str(h):
            completion_idx = i
            break
    if completion_idx is None:
        completion_idx = 9  # fallback
    warranty_idx = completion_idx + 1
    payment_idx = warranty_idx + 1
    sales_idx = payment_idx + 1
    line_idx = sales_idx + 1
    # 電話/傳真在完工前一欄
    phone_fax_idx = completion_idx - 1
    print(f'  欄位偵測: 完工={completion_idx}, 保固={warranty_idx}, 付款={payment_idx}, 業務={sales_idx}')

    for row in ws.iter_rows(min_row=2, max_col=max(line_idx+1, 14), values_only=False):
        col_a = get_cell(row, 0)  # 客戶編號
        col_b = get_cell(row, 1)  # 客戶名稱
        col_d = get_cell(row, 3)  # 地址
        col_e = get_cell(row, 4)  # 聯絡人/電話
        col_f = get_cell(row, 5)  # 手機
        col_g = get_cell(row, 6)  # 手機號碼
        col_i = get_cell(row, phone_fax_idx) if phone_fax_idx < len(row) else None  # 電話/傳真
        col_j = row[completion_idx].value if completion_idx < len(row) else None  # 完工日期
        col_k = get_cell(row, warranty_idx)  # 保固
        col_l = get_cell(row, payment_idx)  # 付款方式
        col_m = get_cell(row, sales_idx)  # 業務
        col_n = get_cell(row, line_idx)  # LINE

        # 跳過空行
        if not col_b and not col_a:
            continue

        # 解析進件分公司和客戶名稱
        source_company, name_after_prefix = extract_source_company(col_b or '')

        # 解析統編和發票資訊
        tax_id, invoice_title, customer_name = extract_invoice_info(name_after_prefix)

        # 也從地址欄找統編
        if not tax_id and col_d:
            tax_id = extract_tax_id(str(col_d))

        # 解析完工日期
        completion_date = parse_roc_date(col_j)

        # 解析保固期限欄：日期 → warranty_date，文字 → note
        warranty_date = None
        warranty_note_text = None
        if col_k:
            wd = parse_roc_date(col_k)
            if wd:
                warranty_date = wd
            else:
                warranty_note_text = str(col_k).strip()

        # 解析聯絡人
        contacts = parse_contacts(col_e, col_f, col_g)

        # 第一位聯絡人作為主要聯絡人
        primary_contact = contacts[0]['name'] if contacts else ''
        primary_phone = contacts[0]['phone'] if contacts else ''

        # 補充電話
        fax = ''
        phone = ''
        if col_i:
            i_str = str(col_i)
            if 'F:' in i_str or 'f:' in i_str:
                fax = i_str.replace('F:', '').replace('f:', '').strip()
            else:
                phone = i_str

        # 自動分類
        category = auto_category(customer_name, col_d)

        record = {
            'original_customer_no': parse_customer_no(col_a),
            'source_company': source_company,
            'customer_name': customer_name,
            'raw_name': col_b,
            'category': category,
            'tax_id': tax_id,
            'site_address': col_d,
            'contact_person': primary_contact,
            'phone': phone or primary_phone,
            'mobile': str(col_g).strip() if col_g else '',
            'fax': fax,
            'contacts': contacts,
            'completion_date': completion_date,
            'warranty_date': warranty_date,
            'warranty_note': warranty_note_text,
            'payment_info': col_l,
            'salesperson_name': col_m,
            'line_official': col_n,
            'source_branch': branch,
        }
        records.append(record)

    print(f'  Parsed: {len(records)} records')
    return records

def assign_numbers(records):
    """產生進件編號和客戶編號"""
    # 按完工日期排序（None 排最後）
    def sort_key(r):
        d = r.get('completion_date')
        if d:
            return d
        return '9999-12-31'

    records.sort(key=sort_key)

    # 產生進件編號：按完工年份分組
    year_counters = {}
    for r in records:
        d = r.get('completion_date')
        if d:
            year = d[:4]
        else:
            year = '2026'  # 無日期的歸到今年
        if year not in year_counters:
            year_counters[year] = 0
        year_counters[year] += 1
        r['case_number'] = f'{year}-{year_counters[year]:04d}'

    # 產生客戶編號：A-六位流水號，按排序順序
    for idx, r in enumerate(records):
        r['customer_no'] = f'A-{idx + 1:06d}'

    return records

def build_groups(records):
    """建立關聯群組"""
    # 方式1: 統編相同 → 同群組
    tax_groups = {}
    # 方式2: 客戶名稱完全相同 → 同群組
    name_groups = {}

    group_id = 0

    for r in records:
        assigned = False

        # 優先用統編關聯
        if r.get('tax_id'):
            tid = r['tax_id']
            if tid not in tax_groups:
                group_id += 1
                tax_groups[tid] = group_id
            r['related_group_id'] = tax_groups[tid]
            assigned = True

        # 用客戶名稱關聯（去除空白後比對）
        if not assigned and r.get('customer_name'):
            cname = r['customer_name'].strip()
            if cname in name_groups:
                r['related_group_id'] = name_groups[cname]
            else:
                # 檢查是否有其他同名的
                pass  # 先標記，等全部掃完再處理

    # 第二輪：客戶名稱重複的建立群組
    name_count = {}
    for r in records:
        if r.get('related_group_id'):
            continue
        cname = (r.get('customer_name') or '').strip()
        if not cname:
            continue
        if cname not in name_count:
            name_count[cname] = []
        name_count[cname].append(r)

    for cname, recs in name_count.items():
        if len(recs) > 1:
            group_id += 1
            for r in recs:
                r['related_group_id'] = group_id

    # 統計
    grouped = sum(1 for r in records if r.get('related_group_id'))
    print(f'  關聯群組: {group_id} 組, 涉及 {grouped} 筆客戶')

    return records, group_id

def main():
    all_records = []

    for f in FILES:
        filepath = f'{EXCEL_DIR}/{f["file"]}'
        try:
            records = process_excel(filepath, f['branch'])
            all_records.extend(records)
        except Exception as e:
            print(f'  Error: {e}')

    print(f'\n總計: {len(all_records)} 筆')

    # 產生編號
    print('\n產生編號...')
    all_records = assign_numbers(all_records)

    # 建立關聯
    print('建立關聯群組...')
    all_records, total_groups = build_groups(all_records)

    # 統計
    with_date = sum(1 for r in all_records if r.get('completion_date'))
    with_tax = sum(1 for r in all_records if r.get('tax_id'))
    with_contacts = sum(1 for r in all_records if r.get('contacts'))

    print(f'\n=== 統計 ===')
    print(f'  總筆數: {len(all_records)}')
    print(f'  有完工日期: {with_date}')
    print(f'  有統編: {with_tax}')
    print(f'  有聯絡人: {with_contacts}')
    print(f'  關聯群組: {total_groups} 組')

    # 分類統計
    cats = {}
    for r in all_records:
        c = r.get('category', 'other')
        cats[c] = cats.get(c, 0) + 1
    print(f'\n客戶分類:')
    for k, v in sorted(cats.items(), key=lambda x: -x[1]):
        label = CATEGORY_LABELS.get(k, k)
        print(f'  {label}({k}): {v}')

    # 進件分公司統計
    src = {}
    for r in all_records:
        s = r.get('source_company') or '(無)'
        src[s] = src.get(s, 0) + 1
    print(f'\n進件分公司:')
    for k, v in sorted(src.items(), key=lambda x: -x[1]):
        print(f'  {k}: {v}')

    # 來源分公司統計
    br = {}
    for r in all_records:
        b = r.get('source_branch') or '(無)'
        br[b] = br.get(b, 0) + 1
    print(f'\n來源:')
    for k, v in br.items():
        print(f'  {k}: {v}')

    # 輸出前5筆預覽
    print(f'\n=== 前5筆預覽 ===')
    for r in all_records[:5]:
        print(f'  {r["customer_no"]} | {r["case_number"]} | {r["original_customer_no"]} | <{r["source_company"]}> {r["customer_name"]} | {r["completion_date"]} | 統編:{r.get("tax_id","-")}')

    # 輸出 JSON
    # 移除不需要的 raw_name
    for r in all_records:
        r.pop('raw_name', None)

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(all_records, f, ensure_ascii=False, indent=None)

    print(f'\n輸出: {OUTPUT_FILE}')
    print(f'檔案大小: {len(json.dumps(all_records, ensure_ascii=False)):,} bytes')

if __name__ == '__main__':
    main()
